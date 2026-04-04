#!/usr/bin/env python3
"""
Export Content Batch to Excel

Parses all markdown content files in a batch folder and generates a formatted
Excel (.xlsx) workbook with one sheet per platform. Designed for easy copy/paste
and quick navigation.

Usage:
    python src/export_content_batch.py outputs/drafts/content-batch-YYYY-MM-DD/
    python src/export_content_batch.py  # auto-detects most recent batch folder
"""

import os
import re
import sys
import glob
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import load_brand_config


def get_brand_colors(config_path=None):
    """Load brand colors from brand_config.json for Excel styling."""
    try:
        config = load_brand_config(config_path)
        return {
            "primary": config.get_color_hex("primary").lstrip("#"),
            "secondary": config.get_color_hex("secondary").lstrip("#"),
            "accent": config.get_color_hex("accent").lstrip("#"),
            "light": config.get_color_hex("neutral_light").lstrip("#"),
            "dark": config.get_color_hex("neutral_dark").lstrip("#"),
            "mid": config.get_color_hex("neutral_mid").lstrip("#"),
        }
    except (FileNotFoundError, Exception):
        return {
            "primary": "2B2B2B",
            "secondary": "5F5F5F",
            "accent": "B08D57",
            "light": "F6F7F5",
            "dark": "1E2428",
            "mid": "9AA3A8",
        }


def parse_content_file(filepath: str) -> dict:
    """Parse a single markdown content file and extract structured fields."""
    with open(filepath, "r", encoding="utf-8") as f:
        text = f.read()

    result = {
        "source_file": os.path.basename(filepath),
        "title": "",
        "type": "",
        "week": "",
        "theme": "",
        "quarterly_ref": "",
        "strategic_context": "",
        "content": "",
        "visual_assets": "",
        "platform": "",
        "week_num": 0,
        "post_num": 0,
    }

    title_match = re.match(r"^#\s+(.+)", text, re.MULTILINE)
    if title_match:
        result["title"] = title_match.group(1).strip()

    meta_patterns = {
        "type": r"\*\*Type:\*\*\s*(.+)",
        "week": r"\*\*Week:\*\*\s*(.+)",
        "theme": r"\*\*Theme:\*\*\s*(.+)",
        "quarterly_ref": r"\*\*Quarterly plan reference:\*\*\s*(.+)",
        "strategic_context": r"\*\*Strategic context:\*\*\s*(.+)",
    }
    for key, pattern in meta_patterns.items():
        match = re.search(pattern, text)
        if match:
            result[key] = match.group(1).strip()

    fname = os.path.basename(filepath).lower()
    if "blog" in fname:
        result["platform"] = "Blog"
    elif "linkedin" in fname:
        result["platform"] = "LinkedIn"
    elif "facebook" in fname:
        result["platform"] = "Facebook"
    elif "podcast" in fname or "youtube" in fname or "video" in fname:
        result["platform"] = "Video"
    elif "clip" in fname:
        result["platform"] = "Clips"
    elif "twitter" in fname or "x-" in fname:
        result["platform"] = "Twitter"
    elif "instagram" in fname or "reel" in fname:
        result["platform"] = "Instagram"
    elif "email" in fname or "newsletter" in fname:
        result["platform"] = "Email"
    elif "tiktok" in fname:
        result["platform"] = "TikTok"
    else:
        result["platform"] = "Other"

    week_match = re.search(r"week-(\d+)", fname)
    if week_match:
        result["week_num"] = int(week_match.group(1))

    post_match = re.search(r"-(\d+)\.md$", fname)
    if post_match:
        result["post_num"] = int(post_match.group(1))

    content_match = re.search(r"## Content\s*\n(.+?)(?=\n---|\n## Quality)", text, re.DOTALL)
    if content_match:
        result["content"] = content_match.group(1).strip()

    visual_match = re.search(r"## Visual Assets\s*\n(.+?)(?=\n---|\n## Content|\n## Clip)", text, re.DOTALL)
    if visual_match:
        result["visual_assets"] = visual_match.group(1).strip()

    return result


def find_batch_folder(path: str = None) -> str:
    """Find the batch folder to export."""
    if path:
        return path

    drafts_dir = os.path.join(os.path.dirname(__file__), "..", "outputs", "drafts")
    if not os.path.exists(drafts_dir):
        print(f"Error: drafts directory not found at {drafts_dir}")
        sys.exit(1)

    batches = sorted(glob.glob(os.path.join(drafts_dir, "content-batch-*")))
    if not batches:
        print("Error: no content-batch-* folders found in outputs/drafts/")
        sys.exit(1)

    return batches[-1]


def export_to_excel(batch_folder: str, config_path: str = None):
    """Parse all markdown files in a batch and export to branded Excel."""
    colors = get_brand_colors(config_path)

    md_files = sorted(glob.glob(os.path.join(batch_folder, "*.md")))
    if not md_files:
        print(f"No .md files found in {batch_folder}")
        return

    pieces = [parse_content_file(f) for f in md_files]

    platforms = {}
    for piece in pieces:
        plat = piece["platform"]
        if plat not in platforms:
            platforms[plat] = []
        platforms[plat].append(piece)

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=colors["primary"], end_color=colors["primary"], fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Arial", size=10, color=colors["dark"])
    body_align = Alignment(vertical="top", wrap_text=True)

    accent_font = Font(name="Arial", size=10, color=colors["accent"], bold=True)
    alt_fill = PatternFill(start_color=colors["light"], end_color=colors["light"], fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color=colors["mid"]),
        right=Side(style="thin", color=colors["mid"]),
        top=Side(style="thin", color=colors["mid"]),
        bottom=Side(style="thin", color=colors["mid"]),
    )

    columns = ["Week", "Title", "Type", "Theme", "Strategic Context", "Content Preview", "Source File"]
    col_widths = [8, 35, 15, 20, 30, 50, 25]

    for plat_name in sorted(platforms.keys()):
        ws = wb.create_sheet(title=plat_name[:31])
        plat_pieces = sorted(platforms[plat_name], key=lambda p: (p["week_num"], p["post_num"]))

        for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, piece in enumerate(plat_pieces, 2):
            content_preview = piece["content"][:300] + "..." if len(piece["content"]) > 300 else piece["content"]

            values = [
                piece["week"],
                piece["title"],
                piece["type"],
                piece["theme"],
                piece["strategic_context"],
                content_preview,
                piece["source_file"],
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border

                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            ws.cell(row=row_idx, column=2).font = accent_font

        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(plat_pieces) + 1}"
        ws.freeze_panes = "A2"

    batch_name = os.path.basename(batch_folder)
    date_part = batch_name.replace("content-batch-", "").replace("-", "")
    output_filename = f"content-batch-summary-{date_part}.xlsx"
    output_path = os.path.join(batch_folder, output_filename)

    wb.save(output_path)
    print(f"Exported {len(pieces)} pieces across {len(platforms)} platforms to:")
    print(f"  {output_path}")


if __name__ == "__main__":
    batch_path = sys.argv[1] if len(sys.argv) > 1 else None
    config_path = sys.argv[2] if len(sys.argv) > 2 else None
    batch_folder = find_batch_folder(batch_path)
    print(f"Exporting batch: {batch_folder}")
    export_to_excel(batch_folder, config_path)
