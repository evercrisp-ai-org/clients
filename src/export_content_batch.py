#!/usr/bin/env python3
"""
Export Content Batch to Excel

Parses all markdown content files in a batch folder and generates a formatted
Excel (.xlsx) workbook with one sheet per platform (Blog, LinkedIn, Facebook,
YouTube). Designed for easy copy/paste and quick navigation.

Usage:
    python src/export_content_batch.py outputs/drafts/content-batch-2026-02-12/
    python src/export_content_batch.py  # auto-detects most recent batch folder
"""

import os
import re
import sys
import glob
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Brand colors (Capable Wealth palette)
# ---------------------------------------------------------------------------
DEEP_MUTED_BLUE = "243A4B"
BLUE_SLATE = "5F7483"
ANTIQUE_GOLD = "B08D57"
OFF_WHITE = "F6F7F5"
CHARCOAL = "1E2428"
WARM_GRAY = "9AA3A8"


# ---------------------------------------------------------------------------
# Markdown parser
# ---------------------------------------------------------------------------

def parse_content_file(filepath: str) -> dict:
    """Parse a single markdown content file and extract structured fields."""
    with open(filepath, "r", encoding="utf-8") as f:
        text = f.read()

    result = {
        "source_file": os.path.basename(filepath),
        "title": "",
        "type": "",
        "week": "",
        "theme": "",
        "quarterly_ref": "",
        "strategic_context": "",
        "content": "",
        "visual_assets": "",
        "platform": "",
        "week_num": 0,
        "post_num": 0,
    }

    # --- Title (first H1) ---
    title_match = re.match(r"^#\s+(.+)", text, re.MULTILINE)
    if title_match:
        result["title"] = title_match.group(1).strip()

    # --- Metadata fields ---
    meta_patterns = {
        "type": r"\*\*Type:\*\*\s*(.+)",
        "week": r"\*\*Week:\*\*\s*(.+)",
        "theme": r"\*\*Theme:\*\*\s*(.+)",
        "quarterly_ref": r"\*\*Quarterly plan reference:\*\*\s*(.+)",
        "strategic_context": r"\*\*Strategic context:\*\*\s*(.+)",
    }
    for key, pattern in meta_patterns.items():
        match = re.search(pattern, text)
        if match:
            result[key] = match.group(1).strip()

    # --- Platform detection from filename ---
    fname = os.path.basename(filepath).lower()
    if "-blog-" in fname:
        result["platform"] = "Blog"
    elif "-linkedin-" in fname:
        result["platform"] = "LinkedIn"
    elif "-facebook-" in fname:
        result["platform"] = "Facebook"
    elif "-youtube-" in fname:
        result["platform"] = "YouTube"

    # --- Week number for sorting ---
    week_match = re.search(r"week-(\d+)", fname)
    if week_match:
        result["week_num"] = int(week_match.group(1))

    # --- Post number for social posts ---
    post_match = re.search(r"(?:linkedin|facebook)-(\d+)", fname)
    if post_match:
        result["post_num"] = int(post_match.group(1))

    # --- Visual Assets section ---
    visual_match = re.search(
        r"## Visual Assets\s*\n(.*?)(?=\n---\s*\n(?!##\s*Visual))",
        text,
        re.DOTALL,
    )
    if visual_match:
        raw_visual = visual_match.group(1).strip()
        # Clean up into a readable summary
        result["visual_assets"] = _clean_visual_assets(raw_visual)

    # --- Content body ---
    # Content is between the last "---" after Visual Assets and "## Quality Checklist"
    result["content"] = _extract_content_body(text)

    return result


def _clean_visual_assets(raw: str) -> str:
    """Condense visual asset block into a readable summary for a spreadsheet cell."""
    sections = []
    current_heading = ""
    current_items = []

    for line in raw.split("\n"):
        line = line.strip()
        if line.startswith("###"):
            if current_heading and current_items:
                sections.append(f"{current_heading}\n" + "\n".join(current_items))
            current_heading = line.lstrip("#").strip()
            current_items = []
        elif line.startswith("- **"):
            # Extract key: value
            kv_match = re.match(r"- \*\*(.+?):\*\*\s*(.*)", line)
            if kv_match:
                current_items.append(f"{kv_match.group(1)}: {kv_match.group(2)}")
        elif line and current_items:
            # Continuation of previous value
            current_items[-1] += " " + line

    if current_heading and current_items:
        sections.append(f"{current_heading}\n" + "\n".join(current_items))

    return "\n\n".join(sections)


def _extract_content_body(text: str) -> str:
    """Extract the main content body from the markdown file.

    Content is everything after the Visual Assets section's closing ---
    and before the ## Quality Checklist heading.

    File structure:
      # Title
      ## Post Metadata
      ---
      ## Visual Assets
      ### Primary Image
      ...
      ---              <-- content starts after this divider
      (content body)
      ---              <-- optional trailing divider
      ## Quality Checklist
    """
    # Strip out the Quality Checklist and everything after it
    parts = re.split(r"\n## Quality Checklist", text, maxsplit=1)
    before_checklist = parts[0]

    # Find the Visual Assets section
    visual_pos = before_checklist.find("## Visual Assets")
    if visual_pos == -1:
        # Fallback: try to find content after metadata's ---
        visual_pos = 0

    remaining = before_checklist[visual_pos:]

    # Find the first --- divider after "## Visual Assets" -- this closes
    # the visual assets block and the content body starts after it
    divider_matches = list(re.finditer(r"\n---\s*\n", remaining))

    if divider_matches:
        # Content starts after the first --- following Visual Assets
        first_divider = divider_matches[0]
        content = remaining[first_divider.end():].strip()
    else:
        content = remaining.strip()

    # Remove trailing "---" that sits right before Quality Checklist
    content = re.sub(r"\n---\s*$", "", content).strip()

    return content


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

def create_workbook(content_items: List[Dict], batch_date: str) -> Workbook:
    """Create a formatted Excel workbook from parsed content items."""
    wb = Workbook()

    # Platform order
    platforms = ["Blog", "LinkedIn", "Facebook", "YouTube"]

    # Column definitions
    columns = [
        ("Week", 22),
        ("Title", 40),
        ("Theme", 28),
        ("Type", 22),
        ("Strategic Context", 55),
        ("Content", 85),
        ("Visual Asset Brief", 65),
        ("Source File", 38),
    ]

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color=OFF_WHITE)
    header_fill = PatternFill(start_color=DEEP_MUTED_BLUE, end_color=DEEP_MUTED_BLUE, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    subheader_fill = PatternFill(start_color=WARM_GRAY, end_color=WARM_GRAY, fill_type="solid")

    body_font = Font(name="Calibri", size=11, color=CHARCOAL)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=11, color=DEEP_MUTED_BLUE)

    gold_font = Font(name="Calibri", bold=True, size=11, color=ANTIQUE_GOLD)

    thin_border = Border(
        bottom=Side(style="thin", color=WARM_GRAY),
        top=Side(style="thin", color=WARM_GRAY),
        left=Side(style="thin", color=WARM_GRAY),
        right=Side(style="thin", color=WARM_GRAY),
    )

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for platform in platforms:
        # Filter items for this platform
        items = [i for i in content_items if i["platform"] == platform]
        if not items:
            continue

        # Sort by week number, then post number
        items.sort(key=lambda x: (x["week_num"], x["post_num"]))

        ws = wb.create_sheet(title=platform)

        # --- Header row ---
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 30

        # --- Data rows ---
        current_week = None
        row_idx = 2

        for item in items:
            # Insert week separator row if week changed
            if item["week_num"] != current_week:
                if current_week is not None:
                    # Add a blank spacer row between weeks
                    row_idx += 1

                # Week header row
                week_label = f"WEEK {item['week_num']}"
                if item["week"]:
                    week_label = item["week"].upper()

                cell = ws.cell(row=row_idx, column=1, value=week_label)
                cell.font = subheader_font
                cell.fill = subheader_fill
                for c in range(1, len(columns) + 1):
                    ws.cell(row=row_idx, column=c).fill = subheader_fill
                    ws.cell(row=row_idx, column=c).border = thin_border
                ws.row_dimensions[row_idx].height = 24
                current_week = item["week_num"]
                row_idx += 1

            # Content row
            row_data = [
                item["week"],
                item["title"],
                item["theme"],
                item["type"],
                item["strategic_context"],
                item["content"],
                item["visual_assets"],
                item["source_file"],
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border

            # Make title column bold with brand color
            ws.cell(row=row_idx, column=2).font = title_font

            # Make week column use gold accent
            ws.cell(row=row_idx, column=1).font = gold_font

            # Set row height based on content length
            content_len = len(item.get("content", ""))
            if content_len > 2000:
                ws.row_dimensions[row_idx].height = 200
            elif content_len > 500:
                ws.row_dimensions[row_idx].height = 120
            else:
                ws.row_dimensions[row_idx].height = 80

            row_idx += 1

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_idx - 1}"

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def find_latest_batch_folder(base_dir: str) -> Optional[str]:
    """Find the most recent content-batch-* folder in the drafts directory."""
    pattern = os.path.join(base_dir, "outputs", "drafts", "content-batch-*")
    folders = sorted(glob.glob(pattern), reverse=True)
    return folders[0] if folders else None


def extract_batch_date(folder_path: str) -> str:
    """Extract the date from a batch folder name and return as YYYYMMDD."""
    folder_name = os.path.basename(folder_path.rstrip("/"))
    date_match = re.search(r"content-batch-(\d{4})-(\d{2})-(\d{2})", folder_name)
    if date_match:
        return f"{date_match.group(1)}{date_match.group(2)}{date_match.group(3)}"
    return "unknown"


def main():
    # Determine batch folder
    if len(sys.argv) > 1:
        batch_folder = sys.argv[1]
    else:
        # Auto-detect from workspace root
        workspace = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        batch_folder = find_latest_batch_folder(workspace)
        if not batch_folder:
            print("Error: No content-batch-* folder found. Provide a path as argument.")
            sys.exit(1)

    batch_folder = os.path.abspath(batch_folder)
    if not os.path.isdir(batch_folder):
        print(f"Error: Directory not found: {batch_folder}")
        sys.exit(1)

    print(f"Processing batch folder: {batch_folder}")

    # Find all markdown files
    md_files = sorted(glob.glob(os.path.join(batch_folder, "*.md")))
    if not md_files:
        print("Error: No .md files found in the batch folder.")
        sys.exit(1)

    print(f"Found {len(md_files)} content files")

    # Parse all files
    content_items = []
    for filepath in md_files:
        try:
            item = parse_content_file(filepath)
            if item["platform"]:
                content_items.append(item)
                print(f"  Parsed: {item['source_file']} -> {item['platform']}")
            else:
                print(f"  Skipped (no platform detected): {os.path.basename(filepath)}")
        except Exception as e:
            print(f"  Error parsing {os.path.basename(filepath)}: {e}")

    if not content_items:
        print("Error: No valid content files parsed.")
        sys.exit(1)

    # Group summary
    platforms = {}
    for item in content_items:
        platforms.setdefault(item["platform"], 0)
        platforms[item["platform"]] += 1
    print(f"\nContent breakdown:")
    for platform, count in sorted(platforms.items()):
        print(f"  {platform}: {count} pieces")

    # Extract batch date for filename
    batch_date = extract_batch_date(batch_folder)

    # Create workbook
    wb = create_workbook(content_items, batch_date)

    # Save
    output_filename = f"content-batch-summary-{batch_date}.xlsx"
    output_path = os.path.join(batch_folder, output_filename)
    wb.save(output_path)
    print(f"\nExported to: {output_path}")


if __name__ == "__main__":
    main()
